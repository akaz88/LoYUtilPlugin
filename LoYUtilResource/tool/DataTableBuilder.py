import codecs
import collections.abc
import glob
import json
import os.path
import pprint
import re
import sys

import openpyxl


point2cood = lambda x, y: "%s%s" % (openpyxl.utils.cell.get_column_letter(x+1), y+1)
align_right = openpyxl.styles.Alignment(horizontal="right")
align_center = openpyxl.styles.Alignment(horizontal="center")
align_left = openpyxl.styles.Alignment(horizontal="left")


#ゲームのテキストテーブルとMODのtidファイルからテキスト辞書を作成
class NamePool:

    def __init__(self, jpdic):
        self.pool = {}
        if not isinstance(jpdic, list):
            jpdic = [jpdic]
        self.read(jpdic)

    #辞書スタイルのアクセサ
    #存在しないキーにはKeyErrorを投げる
    def __getitem__(self, name):
        return self.pool[name]

    def read(self, jpdic):
        for x in jpdic:
            if not os.path.exists(x):
                print("Error: %s is not exist." % x)
                exit(-1)
            if os.path.isdir(x):
                pat = "%s\\**\\*.tid" % x
                self.read(glob.glob(pat, recursive=True))
            elif x.endswith(".tid"):
                self.read_tid(x)
            else:
                self.read_json(x)

    def read_tid(self, tid):
        with open(tid, encoding="utf-8") as f:
            buf = f.read()
        for s in buf.split(","):
            p = s.strip().split(":")
            if len(p) != 2:
                continue
            k,v = p
            if k in self.pool:
                print("[Warning] %s is already in NamePool." % k)
            self.pool[k] = v

    #ゲームのテキストテーブルの処理
    def read_json(self, jsfile):
        js = json.load(open(jsfile, encoding="utf-8"))
        for d in js["data"]["pairs"]:
            #JapaneseTextData.jsonではkeyがUnicodeエスケープされている
            k = codecs.decode(d["key"], "unicode-escape")
            if k in self.pool:
                print("[Warning] %s is already in NamePool." % k)
            #なぜかvalueはエスケープされていない
            self.pool[k] = d["value"]


#セルのデータを読む
#辞書の場合は":"が入っていたり、空文字の場合は""が入っていたりする
def read_cell(cell):
    #print(row)
    v = cell.value
    if isinstance(v, (int, float, bool)):
        return v
    elif ":" in v:
        return dict([v.split(":")])
    elif isinstance(v, str) and v in ('“”', '""'):
        return ""
    else:
        return str(v)


#行の処理
#リストの場合は[x,x,x]、辞書の場合は{"x:x","x:x","x:x"}という形式になっている
def read_row(row):
    if len(row) == 1:
        return read_cell(row[0])
    elif read_cell(row[0]) == "[":
        l = []
        for i in range(1, len(row) - 1):
            if not row[0].value:
                continue
            v = read_cell(row[i])
            if v in ("[", "{"):
                l.append(read_row(row[i:]))
            else:
                l.append(v)
        return l
    elif read_cell(row[0]) == "{":
        d = {}
        for i in range(1, len(row) - 1):
            if not row[0].value:
                continue
            k, v = read_cell(row[i]).popitem()
            #print(i, row[i].value, k, v)
            if v in ("[", "{"):
                #print(k, v)
                d[k] = read_row(row[i:])
            else:
                d[k] = v
        return d
    else:
        print("[Error]: unexcept type.")
        for x in row:
            print(x.value, end=", ")
        print()
        exit(-1)


#データテーブルのxlsxファイルかを超適当に判断する
def is_datatablexlsx(xls_name):
    wb = openpyxl.load_workbook(xls_name)
    if "info" not in wb.sheetnames:
        ws = wb.worksheets[0]
    else:
        ws = wb["info"]
    return True if ws["A1"].value == "m_Name" else False


#src.xlsx->dest/*.*DataTable.jsonの変換
def build_table(src, dest):
    if not os.path.exists(dest) or not os.path.isdir(dest):
        print("output directory %s is not exists." % dest)
        exit(-1)
    #マップのxlsxファイルを開いたらサイレントに閉じる
    if not is_datatablexlsx(src):
        return
    wb = openpyxl.load_workbook(src)
    #ext = wb["info"]["B1"].value + ".json"
    for ws in wb.worksheets:
        if ws.title == "info":
            continue
        d = {}
        for row in ws.rows:
            v = [x for x in filter(lambda x: x.value != None, row[1:])]
            d[row[0].value] = read_row(v)
        if "m_Name" in d:
            ext = d.pop("m_Name") + ".json"
        else:
            ext = wb["info"]["B1"].value + ".json"
        out = os.path.join(dest, "%s.%s" % (ws.title, ext))
        json.dump(d, open(out, "wt", encoding="utf-8"))


#シートの幅を調整
#DngmapConv.pyからコピペ
def adjust_sheet_dimention(ws):
    for col in ws.columns:
        l = len(str(col[0].value))
        for c in col[1:]:
            l = len(str(c.value)) if len(str(c.value)) > l else l
        idx = col[0].column_letter
        sz = col[0].font.size / 10.0
        #1だけ余白を開けておかないとセルによっては幅の調整がいまいちな感じになる
        ws.column_dimensions[idx].width = l * sz + (1 if l else 0)


#再帰的に行を埋めていく
def write_row(ws, y, x, v):
    if isinstance(v, (int, float, bool)):
        ws[point2cood(x, y)] = v
        ws[point2cood(x, y)].alignment = align_center
        return x + 1
    elif isinstance(v, str):
        if not v:
            ws[point2cood(x, y)] = '""'
            ws[point2cood(x, y)].alignment = align_center
        else:
            ws[point2cood(x, y)] = v
            ws[point2cood(x, y)].alignment = align_center
        return x + 1
    elif isinstance(v, list):
        ws[point2cood(x, y)] = "["
        ws[point2cood(x, y)].alignment = align_right
        x += 1
        for e in v:
            x = write_row(ws, y, x, e)
        ws[point2cood(x, y)] = "]"
        ws[point2cood(x, y)].alignment = align_left
        return x + 1
    elif isinstance(v, dict):
        ws[point2cood(x, y)] = "{"
        ws[point2cood(x, y)].alignment = align_right
        x += 1
        for e in v:
            if isinstance(v[e], (list, dict)):
                ws[point2cood(x, y)] = "%s:" % e
                ws[point2cood(x, y)].alignment = align_center
                x = write_row(ws, y, x, e)
            else:
                ws[point2cood(x, y)] = "%s:%s" % (e, v[e])
                ws[point2cood(x, y)].alignment = align_center
                x += 1
        ws[point2cood(x, y)] = "}"
        ws[point2cood(x, y)].alignment = align_left
        return x + 1


#SrcDataTable.json->dest.xlsxへの変換を行う
#ここでソースに指定するデータテーブルはゲーム本体のものを使用しなければならない
#jpdicにはゲーム本体のJapaneseDataTable.jsonとMODのtidファイルかMODディレクトリを指定する
#jpdicを使用することによりワークシート名にIDではなくゲーム内の名前が表示できる
def build_xlsx(src, dest, jpdic, force=False):
    if not force and os.path.exists(dest):
        print("output file %s is already exists." % dest)
        exit(-1)
    np = NamePool(jpdic)
    js = json.load(open(src, encoding="utf-8"))
    if "records" not in js or "m_Name" not in js:
        print("invalid input format.")
        exit(-1)
    wb = openpyxl.Workbook()
    wb.remove(wb.worksheets[0])
    wb.create_sheet("info")
    wb["info"]["A1"] = "m_Name"
    wb["info"]["A1"].alignment = align_center
    m_Name = js["m_Name"] if not js["m_Name"].endswith("Table") else js["m_Name"][:-5]
    wb["info"]["B1"] = m_Name
    wb["info"]["B1"].alignment = align_center
    adjust_sheet_dimention(wb["info"])
    for d in js["records"]:
        #popitemは末尾からpopしていくので結果をソートしたいなら逆順にしておく必要がある
        d = dict(sorted(d.items(), reverse=True))
        if "name" in d and "id" in d["name"]:
            try:
                ws = wb.create_sheet(np[d["name"]["id"]])
            except KeyError:
                ws = wb.create_sheet("ID：%s" % d["id"])
        elif "id" in d:
            ws = wb.create_sheet("ID：%s" % d["id"])
        else:
            print("[Error] no name or id in json.")
            exit(-1)
        for y in range(0, len(d)):
            #IDを最初の行にもってくるようにする
            if "id" in d:
                k = "id"
                v = d.pop(k)
            else:
                k, v = d.popitem()
            x = 0
            ws[point2cood(x, y)] = k
            ws[point2cood(x, y)].alignment = align_left
            write_row(ws, y, x+1, v)
        adjust_sheet_dimention(ws)
    wb.save(dest)

if __name__ == "__main__":
    if len(sys.argv) < 3 or sys.argv[1] in ("-h", "--help", "/?"):
        print("DataTableBuilder.py <in> <out> [JapaneseTextData.json] [tid file] [mod dir] [-f/--force]")
        exit()
    if sys.argv[1] == sys.argv[2]:
        print("input and output are same file.")
        exit()
    force = False
    if sys.argv[-1] in ("-f", "--force"):
        force = True
        sys.argv.pop(-1)
    if sys.argv[1].endswith(".xlsx"):
        build_table(sys.argv[1], sys.argv[2])
    else:
        build_xlsx(sys.argv[1], sys.argv[2], sys.argv[3:], force=force)