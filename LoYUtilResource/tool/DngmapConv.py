import json
import os.path
import pprint
import re
import sys

import openpyxl


#既定のアセットパッケージリスト
DEFINED_ASSET_PACKAGES = ["Castle_01_Package", "Castle_02_Package", "Castle_03_Package", "Castle_04_Package", "Castle_05_Package", "Castle_06_Package", "Forest_01_Package", "Forest_02_Package", "Forest_03_Package", "Forest_04_Package", "Forest_05_Package", "Forest_06_Package", "Shrine_01_Package", "Shrine_04_Package", "Tunnel_01_Package", "Tunnel_05_Package"]
#各パッケージごとの既定オブジェクトとセル上での記号の対応
DEFINED_OBJECTS_DICT = {
    "Castle": {
            "扉": "Door", "鍵": "ControlledDoor", "障": "Obstacle",
            "⇒": "OnewayDoor", "鳥": "Bird", "柱": "Pillar1",
            "燭": "Candlestick", "毒": "PoisonZone", "ス": "Script",
            "麻": "ParalyzeZone", "沈": "SilenceZone",
            "ダ": "DamageFloor", "→": "ShiftZone", "←": "ShiftZone",
            "↑": "ShiftZone", "↓": "ShiftZone", "転": "Teleporter",
            "上": "UpStairs", "天": "CeilingHole", "穴": "Chute",
                #CeilingHole: 上方向にはしご設置可能
            "下": "DownStairs", "地": "FloorHole",
                #FloorHole: 下方向にはしご設置可能
            "壁": "WallScript", "敵": "Enemy", "宝": "Item",
            "音": "MusicChanger", "": "Fence", "柱": "Pillar2",
            "出": "Exit"
                #Pillar2はセルのフォント設定がイタリック体
        },
    "Tunnel": {
            "扉": "Door", "鍵": "ControlledDoor", "宝": "Item",
            "⇒": "OnewayDoor", "-": "FenceCenter", "柱": "Pillar",
            "手": "Hand", "灯": "Emissive", "毒": "PoisonZone",
            "麻": "ParalyzeZone", "沈": "SilenceZone",
            "透": "DamageFloorInvisible", "ダ": "DamageFloor",
            "→": "ShiftZone", "←": "ShiftZone", "ス": "Script",
            "↑": "ShiftZone", "↓": "ShiftZone", "転": "Teleporter",
            "上": "UpStairs", "天": "CeilingHole", "穴": "Chute",
            "下": "DownStairs", "地": "FloorHole",
            "壁": "WallScript", "敵": "Enemy", "障": "Crucifixion",
            "音": "MusicChanger", "": "Fence", "出": "Exit"
        },
    "Shrine": {
            "扉": "Door", "鍵": "ControlledDoor", "碑": "Obelisk",
            "⇒": "OnewayDoor", "柱": "Pillar", "灯": "Light",
            "毒": "PoisonZone", "麻": "ParalyzeZone",
            "沈": "SilenceZone", "ダ": "DamageFloor",
            "→": "ShiftZone", "←": "ShiftZone", "ス": "Script",
            "↑": "ShiftZone", "↓": "ShiftZone", "転": "Teleporter",
            "上": "UpStairs", "天": "CeilingHole", "穴": "Chute",
            "下": "DownStairs", "地": "FloorHole",
            "壁": "WallScript", "敵": "Enemy", "宝": "Item",
            "音": "MusicChanger", "": "Fence", "出": "Exit"
        },
    "Forest": {
            "扉": "Door", "鍵": "ControlledDoor", "灯": "Emissive",
            "墓": "GraveMarker", "⇒": "OnewayDoor", "柱": "Pillar1",
            "木": "Tree", "透": "DamageFloorInvisible",
            "毒": "PoisonZone", "麻": "ParalyzeZone",
            "沈": "SilenceZone", "ダ": "DamageFloor",
            "→": "ShiftZone", "←": "ShiftZone", "ス": "Script",
            "↑": "ShiftZone", "↓": "ShiftZone", "転": "Teleporter",
            "上": "UpStairs", "天": "CeilingHole", "穴": "Chute",
            "下": "DownStairs", "地": "FloorHole",
            "壁": "WallScript", "敵": "Enemy", "宝": "Item",
            "音": "MusicChanger", "": "Fence", "柱": "Pillar2",
            "出": "Exit"
        }
}
#セルに書いてあっても無視するオブジェクトのリスト
IGNORE_OBJECT_LIST = ["ControlledDoor", "Teleporter", "Script", "UpStairs", "CeilingHole", "Chute", "DownStairs", "FloorHole", "WallScript", "Enemy", "Item", "MusicChanger"]
#マップのサイズはゲーム内で既定なのでそのまま使う
WIDTH = 30
HEIGHT = 30
#セル名縦列から数字への変換テーブル
cell_tbl = {"A": 0, "B": 1, "C": 2, "D": 3, "E": 4, "F": 5, "G": 6, "H": 7, "I": 8, "J": 9, "K": 10, "L": 11, "M": 12, "N": 13, "O": 14, "P": 15, "Q": 16, "R": 17, "S": 18, "T": 19, "U": 20, "V": 21, "W": 22, "X": 23, "Y": 24, "Z": 25, "AA": 26, "AB": 27, "AC": 28, "AD": 29}
#数字からセル名への逆変換テーブル
cell_rtbl = {v: k for k, v in cell_tbl.items()}
re_splitter = re.compile("(?P<x>[A-Z]+)(?P<y>[0-9]+)")
#床タイプ(黄: 床, 青: 浅瀬, 緑: 深瀬, 黒: 奈落, 白: 壁)
tiletype_tbl = {"FFFFFF00": 0, "FF0000FF": 1, "FF00FF00": 2, "FF000000": 3, "FFFFFFFF": 4}
tiletype_rtbl = {v: k for k, v in tiletype_tbl.items()}
#Point(int, int)からblocksシートのブロックセル名への変換
pointdict2coodinate = lambda d: "%s%s" % (cell_rtbl[d["x"]], (HEIGHT - d["y"]))
#(int, int)からセル名への変換
point2coodinate = lambda x, y: "%s%s" % (openpyxl.utils.cell.get_column_letter(x+1), y+1)


#mapObjectEntryを作成
def create_object(d, name, point, facing, components=[]):
    if "_" not in name:
        name = "%s_%s%s" % (name, point, facing)
    #if name.split("_")[0] not in DEFINED_OBJECTS:
    defined = DEFINED_OBJECTS_DICT[d["assetPackage"]["package"].split("_")[0]].values()
    if name.split("_")[0] not in defined:
        print("[create_object]Error: %s is not valid object name." % name)
        exit(-1)
    point = re_splitter.fullmatch(point).groupdict()
    point["x"] = int(cell_tbl[point["x"]])
    point["y"] = 30 - int(point["y"])
    return {"name": name, "point": point, "facing": int(facing), "components": components}


#フェンスをmapObjectEntriesに追加
def add_fence(cell, d):
    l = [cell.border.top, cell.border.right, cell.border.bottom, cell.border.left]
    for i in range(0, 4):
        if l[i] is None or l[i].style is None:
            continue
        d["mapObjectEntries"].append(create_object(d, "Fence", cell.coordinate, i))


#セルに書かれた値を元にオブジェクトを生成してmapObjectEntriesに追加
def add_object(cell, d):
    asset_type = d["assetPackage"]["package"].split("_")[0]
    obj_dict = DEFINED_OBJECTS_DICT[asset_type]
    name = obj_dict[cell.value[0]]
    point = cell.coordinate
    facing = int(cell.value[1])
    #Pillar1/2の別があるのはCastleとForestのみ
    if name.startswith("Pillar") and asset_type in ("Castle", "Forest"):
        if cell.font.i:
            name = "Pillar2"
        else:
            name = "Pillar1"
    #詳細に設定しなければならないオブジェクトはセルに書いてあっても無視する
    if name not in IGNORE_OBJECT_LIST:
        d["mapObjectEntries"].append(create_object(d, name, point, facing))
    #一つのセルに複数のオブジェクトが入っていたらvalueを変更してもう一度回す
    if(cell.value > 2):
        cell.value = cell.value[2:]
        add_object(cell, d)


#objectsシートからmapObjectEntriesを作成
def read_objects(ws, d):
    for row in ws.rows:
        if not row[0].value:
            continue
        name = row[0].value
        point = row[1].value
        facing = row[2].value
        components = []
        if len(row) > 3 and row[3].value:
            key = ("name", "value")
            for x in row[3:]:
                if x.value and ":" in x.value:
                    components.append(dict(zip(key, x.value.split(":"))))
        d["mapObjectEntries"].append(create_object(d, name, point, facing, components))


#blocksシートからブロックデータとフェンス等のオブジェクトを読む
def read_blocks(ws, d):
    defined = DEFINED_OBJECTS_DICT[d["assetPackage"]["package"].split("_")[0]].values()
    #Excelの縦行のカウントとゲームでの縦行のカウントはスタート位置が逆
    for y in reversed(range(1, HEIGHT + 1)):
        #print(ws[y], len(ws[y]))
        for x in range(0, WIDTH):
            cell = ws[y][x]
            b = {"TileType": 0, "TileVariationId": 0, "Attributes": 0, "BattleStageType": 0, "SurfaceType": 0}
            #print("%s, %s" % (cell.coordinate, cell.fill.bgColor.rgb))
            b["TileType"] = tiletype_tbl[cell.fill.fgColor.rgb]
            d["blocks"].append(b)
            if cell.border.left or cell.border.right or cell.border.top or cell.border.bottom:
                add_fence(cell, d)
            if cell.value and cell.value[0] in defined:
                add_object(cell, d)


#infoシートからデータを読む
def read_info(ws, d):
    d["mapName"] = ws["A1"].value
    d["assetPackage"] = {"package": ws["A2"].value}
    d["width"] = WIDTH
    d["height"] = HEIGHT
    d["table"] = []


#マップのxlsxファイルかを超適当に判断する
def is_dngmapxlsx(xls_name):
    wb = openpyxl.load_workbook(xls_name)
    try:
        return True if wb["info"]["A1"].value != "m_Name" else False
    except KeyError:
        return False


#LibreOfficeで書いたXLSXファイルをdngmap.jsonに変換
def xls2json(xls_name, json_name, force=False):
    if not force and os.path.exists(json_name):
        print("output file %s is already exists." % json_name)
        exit(-1)
    #print("convert %s to %s" % (xls_name, json_name))
    #データテーブルのxlsxファイルを開いたらサイレントに閉じる
    if not is_dngmapxlsx(xls_name):
        return
    wb = openpyxl.load_workbook(xls_name)
    d = {}
    d["blocks"] = []
    d["mapObjectEntries"] = []
    read_info(wb["info"], d)
    #pprint.pp(d)
    read_blocks(wb["blocks"], d)
    read_objects(wb["objects"], d)
    #pprint.pp(d["mapObjectEntries"])
    with open(json_name, "wt") as f:
        json.dump(d, f)


#シートの幅を調整
def adjust_sheet_dimention(ws):
    for col in ws.columns:
        l = len(str(col[0].value))
        for c in col[1:]:
            l = len(str(c.value)) if len(str(c.value)) > l else l
        idx = col[0].column_letter
        sz = col[0].font.size / 10.0
        #1だけ余白を開けておかないとセルによっては幅の調整がいまいちな感じになる
        ws.column_dimensions[idx].width = l * sz + (1 if l else 0)


#objectsシートにマップオブジェクト（除くblocksシートに描く分）を書いていく
def create_objects_sheet(ws, d):
    i = 0
    for obj in sorted(d["mapObjectEntries"], key=lambda x: x["name"]):
        #Fence等は自動でmapObjectEntriesから削除されるようになった
        #if obj["name"].startswith("Fence") or :
            #continue
        ws[point2coodinate(0, i)].value = obj["name"]
        ws[point2coodinate(1, i)].value = pointdict2coodinate(obj["point"])
        ws[point2coodinate(2, i)].value = obj["facing"]
        if obj["components"]:
            j = 3
            for x in obj["components"]:
                if x["name"] == "Destination":
                    point = {k: int(v) for k, v in zip(["x", "y"], x["value"].split(","))}
                    ws[point2coodinate(j, i)].value = "%s:%s" % (x["name"], pointdict2coodinate(point))
                else:
                    ws[point2coodinate(j, i)].value = "%s:%s" % (x["name"], x["value"])
                j += 1
        i += 1


#blocksシートにフェンスを描く
def add_fence_block(ws, obj):
    point = pointdict2coodinate(obj["point"])
    facing = obj["facing"]
    #print(obj["point"], point, facing)
    color = openpyxl.styles.colors.Color(rgb="FF888888")
    side = openpyxl.styles.borders.Side(style="thick", color=color)
    border = ws[point].border
    #top
    if facing == 0:
        ws[point].border = openpyxl.styles.borders.Border(top=side, right=border.right, bottom=border.bottom, left=border.left, diagonal=border.diagonal, diagonalUp=border.diagonalUp, diagonalDown=border.diagonalDown)
    #right
    elif facing == 1:
        ws[point].border = openpyxl.styles.borders.Border(top=border.top, right=side, bottom=border.bottom, left=border.left, diagonal=border.diagonal, diagonalUp=border.diagonalUp, diagonalDown=border.diagonalDown)
    #bottom
    elif facing == 2:
        ws[point].border = openpyxl.styles.borders.Border(top=border.top, right=border.right, bottom=side, left=border.left, diagonal=border.diagonal, diagonalUp=border.diagonalUp, diagonalDown=border.diagonalDown)
    #left
    elif facing == 3:
        ws[point].border = openpyxl.styles.borders.Border(top=border.top, right=border.right, bottom=border.bottom, left=side, diagonal=border.diagonal, diagonalUp=border.diagonalUp, diagonalDown=border.diagonalDown)


#[Obsolete]: blocksシートに柱を描く：／ならPillar1、＼ならPillar2
#blocksシートに柱を描く：普通の字体ならPillar1、イタリックならPillar2
def add_pillar_block(ws, obj):
    name = obj["name"]
    point = pointdict2coodinate(obj["point"])
    if name.startswith("Pillar2") and asset_type in ("Castle", "Forest"):
        ws[point].font = openpyxl.styles.fonts.Font(i=True)
    ws[point].value = "柱0"
    return

    name = obj["name"]
    point = pointdict2coodinate(obj["point"])
    print(obj["point"], point, facing)
    color = openpyxl.styles.colors.Color(rgb="FF880088")
    side = openpyxl.styles.borders.Side(style="thick", color=color)
    border = ws[point].border
    if name.startswith("Pillar1"):
        ws[point].border =  openpyxl.styles.borders.Border(top=border.top, right=border.right, bottom=border.bottom, left=border.left, diagonal=side, diagonalUp=True, diagonalDown=border.diagonalDown)
    else:
        ws[point].border =  openpyxl.styles.borders.Border(top=border.top, right=border.right, bottom=border.bottom, left=border.left, diagonal=side, diagonalUp=border.diagonalUp, diagonalDown=True)


#blocksシートのセルにオブジェクトを書き込む
def add_object_in_cell(ws, symbol, obj):
    point = pointdict2coodinate(obj["point"])
    facing = obj["facing"]
    value = "%s%s" % (symbol, facing)
    if len(value) > 2:
        print("[add_object_in_cell]Error: %s@%s is not valid symbol." % (value, point))
        exit(-1)
    if ws[point].value:
        ws[point].value += value
    else:
        ws[point].value = value


#blocksシートに書き込む必要のあるオブジェクトを変換
def add_block_objects(wb, d):
    blocks = wb["blocks"]
    objects = wb["objects"]
    defined = DEFINED_OBJECTS_DICT[d["assetPackage"]["package"].split("_")[0]]
    rdefined = {v: k for k, v in defined.items()}
    del_idx = []
    for obj in d["mapObjectEntries"]:
        name = obj["name"].split("_")[0]
        if name == "Fence":
            add_fence_block(blocks, obj)
        #Pillar1/2:facingにかかわらず常にブロック右上に配置
        elif name.startswith("Pillar"):
            add_pillar_block(blocks, obj)
        elif name not in rdefined:
            print("[add_block_objects]Error: %s is not valid object name." % name)
            exit(-1)
        else:
            #obj["name"] = rdefined[name]
            add_object_in_cell(blocks, rdefined[name], obj)
        if name not in IGNORE_OBJECT_LIST:
            del_idx.append(d["mapObjectEntries"].index(obj))
    del_idx.sort(reverse=True)
    for i in del_idx:
        del d["mapObjectEntries"][i]


#タイルブロックをblocksシートに変換
def create_blocks_sheet(ws, d):
    #幅と高さを適当に調整@LibreOffice
    for k in cell_tbl.values():
        ws.row_dimensions[k].height = 15
    for k in cell_tbl.keys():
        ws.column_dimensions[k].width = 2.7
    for y in range(0, HEIGHT):
        #print(cell_rtbl[y-1], ":", end="")
        for x in range(0, WIDTH):
            rgb = tiletype_rtbl[d["blocks"][(HEIGHT - y - 1) * HEIGHT + x]["TileType"]]
            ws[point2coodinate(x, y)].fill = openpyxl.styles.fills.PatternFill(patternType=openpyxl.styles.fills.FILL_SOLID, fgColor=openpyxl.styles.colors.Color(rgb=rgb))
            #ws[point2coodinate(x, y)].fill.fgColor.rgb = ws[point2coodinate(x, y)].fill.bgColor.rgb = rgb


#infoシートを作成
def create_info_sheet(ws, d):
    ws["A1"] = d["mapName"]
    try:
        ws["A2"] = d["assetPackage"]["package"]
    except KeyError:
        if d["mapName"].startswith("Castle"):
            d["assetPackage"]["package"] = "Castle_01_Package"
        elif d["mapName"].startswith("Tunnel"):
            d["assetPackage"]["package"] = "Tunnel_01_Package"
        elif d["mapName"].startswith("Shrine"):
            d["assetPackage"]["package"] = "Shrine_01_Package"
        elif d["mapName"].startswith("Forest"):
            d["assetPackage"]["package"] = "Forest_01_Package"
        print("[create_info_sheet]cannnot find valid asset package. use alternative package %s." % d["assetPackage"]["package"])
        ws["A2"] = d["assetPackage"]["package"]


#json->xlsxの変換を行う
def json2xlsx(json_name, xls_name, force=False):
    if not force and os.path.exists(xls_name):
        print("output file %s is already exists." % xls_name)
        exit(-1)
    #print("convert %s to %s" % (json_name, xls_name))
    d = json.load(open(json_name, encoding="utf8"))
    wb = openpyxl.Workbook()
    #デフォルトで作成されるシートを除去
    wb.remove(wb.worksheets[0])
    wb.create_sheet("info")
    wb.create_sheet("blocks")
    wb.create_sheet("objects")
    create_info_sheet(wb["info"], d)
    create_blocks_sheet(wb["blocks"], d)
    add_block_objects(wb, d)
    create_objects_sheet(wb["objects"], d)
    adjust_sheet_dimention(wb["info"])
    adjust_sheet_dimention(wb["objects"])
    wb.save(xls_name)
    #print("done")


if __name__ == "__main__":
    if len(sys.argv) < 4 or sys.argv[1] in ("-h", "--help", "/?") or sys.argv[3][3:] not in ("encode", "decode"):
        print("DngmapConv.py <in> <out> -t=decode/encode [-f/--force]")
        exit()
    if sys.argv[1] == sys.argv[2]:
        print("input and output are same file.")
        exit()
    force = False
    if len(sys.argv) > 4:
        for a in sys.argv[4:]:
            if a in ("-f", "--force"):
                force = True
            else:
                print("invalid argument : %s" % a)
                exit()
    if sys.argv[3][3:] == "encode":
        json2xlsx(sys.argv[1], sys.argv[2], force=force)
    else:
        xls2json(sys.argv[1], sys.argv[2], force=force)
